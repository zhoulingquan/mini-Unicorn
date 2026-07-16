#!/usr/bin/env python3
"""用 openpyxl 创建 Excel 文档。

用法:
    # 基本表格
    python create_xlsx.py output.xlsx --sheet "Sheet1" --data '[["姓名","年龄"],["张三",25],["李四",30]]'

    # 多 sheet
    python create_xlsx.py output.xlsx --sheet "Q1" "Q2" --data '[["营收"],["1.2亿"]]' '[["营收"],["1.5亿"]]'

    # 带格式(标题行加粗)
    python create_xlsx.py output.xlsx --sheet "员工" --data '[["姓名","部门","职级"],["张三","研发","P6"]]' --bold-header
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


def create_xlsx(
    output_path: Path,
    *,
    sheets: list[str],
    data_blocks: list[str],
    bold_header: bool = False,
) -> None:
    """创建 .xlsx 文档。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("错误:需要 openpyxl 库。安装:pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if len(sheets) != len(data_blocks):
        print(f"错误:sheet 数量({len(sheets)})与 data 数量({len(data_blocks)})不匹配", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    # 移除默认 sheet
    wb.remove(wb.active)

    for sheet_name, data_json in zip(sheets, data_blocks):
        try:
            rows = json.loads(data_json)
        except json.JSONDecodeError as e:
            print(f"错误:sheet '{sheet_name}' 的 data JSON 解析失败: {e}", file=sys.stderr)
            sys.exit(1)

        if not isinstance(rows, list):
            print(f"错误:sheet '{sheet_name}' 的 data 必须是二维数组", file=sys.stderr)
            sys.exit(1)

        ws = wb.create_sheet(title=sheet_name)

        for i, row in enumerate(rows, 1):
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=val)
                if bold_header and i == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")

        # 自动列宽(简单估算)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"已创建: {output_path} ({len(sheets)} sheet)")


def main() -> None:
    parser = argparse.ArgumentParser(description="创建 Excel 文档")
    parser.add_argument("output", help="输出 .xlsx 路径")
    parser.add_argument("--sheet", nargs="+", required=True, help="sheet 名称(可多个)")
    parser.add_argument(
        "--data",
        nargs="+",
        required=True,
        help='表格数据 JSON,如 \'[["A","B"],["1","2"]]\',每个 --sheet 对应一个 --data',
    )
    parser.add_argument("--bold-header", action="store_true", help="标题行加粗+居中+底色")
    args = parser.parse_args()

    create_xlsx(
        Path(args.output).resolve(),
        sheets=args.sheet,
        data_blocks=args.data,
        bold_header=args.bold_header,
    )


if __name__ == "__main__":
    main()
